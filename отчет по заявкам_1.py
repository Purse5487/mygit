import openpyxl
import re

# Открываем файл Excel
workbook = openpyxl.load_workbook('113.xlsx')

# Выбираем активный лист (может потребоваться изменить имя листа)
sheet = workbook.active

# Функция для обработки текста
def process_text(text):
    match = re.search(r'№ 8629_0\d{4}', text)
    if match:
        return match.group(0)
    elif "Административное здание г Великий Новгород  пр-кт Мира 32к1" in text:
        return "Здание"
    elif "Кассово-инкасаторский центр г Боровичи  ул Алексея Кузнецова 38" in text:
        return "№ 8629_01850"
    elif "Административное здание г Великий Новгород  пр-кт Мира 44/20" in text:
        return "Гараж"
    elif "ВСП г Великий Новгород  ул Людогоща 6/13" in text:
        return "№ 8629_01420"
    elif "ж/д_ст Уторгош ул Пионерская 65" in text:
        return "№ 8629_01532"
    else:
        return text

# Перебираем все ячейки в столбце bs (столбец 71)
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=71, max_col=71):
    cell = row[0]  # Выбираем текущую ячейку в столбце bs
    text = cell.value  # Получаем значение из текущей ячейки

    if text:  # Проверяем, не является ли значение пустым
        processed_text = process_text(text)  # Применяем функцию process_text к значению
        # Записываем результат в столбец cf (столбец 84)
        sheet.cell(row=cell.row, column=84, value=processed_text)

# Сохраняем изменения в файле Excel
workbook.save('113_updated.xlsx')

