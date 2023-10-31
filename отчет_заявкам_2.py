# Заполнение столбца CD шаблонами работ
import openpyxl
import tkinter as tk
from tkinter import filedialog

# Создаем графическое окно для выбора файла
root = tk.Tk()
root.withdraw()  # Скрываем главное окно

# Запрос пути к файлу с помощью диалогового окна
file_path = filedialog.askopenfilename(title="Выберите файл Excel", filetypes=[("Excel Files", "*.xls *.xlsx")])

if not file_path:
    print("Файл не выбран. Программа завершена.")
else:
    # Открываем Excel-файл
    wb = openpyxl.load_workbook(file_path)

    # Выбираем активный лист
    sheet = wb.active

# Функция для обработки текста и записи в столбец cd
    def process_bf(text, bs_text):
        if "Монтаж/демонтаж аксессуаров интерьераВид декоративного элемента:" in text:
            return "Монтаж/демонтаж аксессуаров интерьера"
        elif "Монтаж/демонтаж мебелиТип мебели:" in text:
            return "Монтаж/демонтаж мебели"
        elif "Ремонт мебели и аксессуаров интерьера" in text:
            if "Ремонт аксессуаров интерьераСпособ заполнения:" in bs_text:
                return "Ремонт аксессуаров интерьера"
            elif "Ремонт мебелиСпособ заполнения:" in bs_text:
                return "Ремонт мебели"
            elif "Вскрытие запирающего устройства" in bs_text:
                return "Ремонт мебели"
            else:
                return "Значение не найдено"
        else:
            return text

    # Обрабатываем данные в столбце bf и записываем результаты в столбец cd
    for row_number in range(2, sheet.max_row + 1):
        bf_value = sheet.cell(row=row_number, column=58).value
        bs_value = sheet.cell(row=row_number, column=71).value
        cd_value = process_bf(bf_value, bs_value)
        sheet.cell(row=row_number, column=82, value=cd_value)

    # Сохраняем обновленный Excel-файл
    output_file_path = "113_updated.xlsx"
    wb.save(output_file_path)
    print(f"Данные обработаны и сохранены в файл: {output_file_path}")
