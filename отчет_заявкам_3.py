from openpyxl import load_workbook
from datetime import datetime

# Загрузка файлов Excel
book = load_workbook(filename="113.xlsx")
sheet = book.active
book1 = load_workbook(filename='114.xlsx')
sheet1 = book1.active

# Начальные значения
i1 = 850
z4 = datetime(2023, 9, 1, 0, 0, 0)
z5 = datetime(2023, 9, 30, 23, 59, 59)
z6 = "отозвано"
i2 = 19

# Цикл обработки строк
for i in range(2, i1):
    a = sheet['a' + str(i)].value
    cd = sheet['cd' + str(i)].value
    cf = sheet['cf' + str(i)].value
    j = sheet['j' + str(i)].value
    date_z = sheet['x' + str(i)].value  # Получаем значение даты
    if date_z == '-':
        pass
    else:
        # Проверяем, что значение даты не является None
        if date_z is not None:
            if z4 <= date_z <= z5 and j != z6:
                sheet1['a' + str(i2)] = a
                sheet1['b' + str(i2)] = cf
                sheet1['d' + str(i2)] = cd
                i2 += 1

# Сохранение и закрытие файлов
book.save("113.xlsx")
book.close()
book1.save("114.xlsx")
book1.close()

print("Файл обработан")
