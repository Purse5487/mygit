import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook

def process_excel_files():
    # Открываем диалоговые окна для выбора двух Excel файлов
    file_path1 = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")], title="Выберите первый файл Excel")
    file_path2 = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")], title="Выберите второй файл Excel")

    if file_path1 and file_path2:
        try:
            # Загружаем первый и второй файлы
            workbook1 = load_workbook(file_path1)
            workbook2 = load_workbook(file_path2)

            # Получаем активные листы из обоих файлов
            worksheet1 = workbook1.active
            worksheet2 = workbook2.active

            # Инициализируем счетчик для подсчета совершенных записей
            count = 0

            # Проходим по всем строкам первого файла
            for row1 in worksheet1.iter_rows(values_only=True):
                cell_value1_col1 = row1[0]  # Значение в 1 столбце (A)
                cell_value1_col6 = row1[5]  # Значение в 6 столбце (F)
                cell_value1_col12 = row1[11]  # Значение в 12 столбце (L)

                # Если 12 столбец пуст и условия совпадения выполняются
                if not cell_value1_col12 and cell_value1_col1 == 4600007963 and cell_value1_col6 in ["В работе", "Новая"]:
                    # Проходим по всем строкам второго файла
                    for row2 in worksheet2.iter_rows(min_col=2, max_col=2, values_only=True):
                        cell_value2_col2 = row2[0]  # Значение во 2 столбце (B)

                        # Если найдено совпадение во втором файле
                        if cell_value2_col2 == cell_value1_col1:
                            # Находим и извлекаем число после "Объем работ: "
                            cell_value2_col6 = row2[1]  # Значение в 6 столбце (F)
                            volume_str = cell_value2_col6.split("Объем работ: ")[1].split()[0]

                            # Записываем найденное значение в 12 столбец первого файла
                            row1[11] = volume_str
                            count += 1

            # Сохраняем изменения в первом файле
            workbook1.save(file_path1)

            if count > 0:
                print(f"Выполнено {count} записей.")
            else:
                print("Записей не найдено.")

        except Exception as e:
            # Обрабатываем возможные ошибки при обработке файлов
            print(f"Ошибка при обработке файлов: {str(e)}")
    else:
        # В случае, если файлы не выбраны
        print("Не выбраны оба файла.")

# Создаем графическое окно
root = tk.Tk()
root.title("Обработка двух Excel файлов")

# Создаем кнопку для выбора файлов и вызываем функцию process_excel_files при ее нажатии
select_button = tk.Button(root, text="Выбрать Excel файлы", command=process_excel_files)
select_button.pack()

# Запускаем главный цикл Tkinter
root.mainloop()
