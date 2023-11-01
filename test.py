import tkinter as tk
from tkinter import filedialog
import openpyxl

# Определяем функцию для обработки Excel-файлов
def process_excel_files():
    # Получаем пути к выбранным файлам из текстовых полей
    file1_path = file1_entry.get()
    file2_path = file2_entry.get()

    try:
        # Загружаем оба файла Excel с помощью openpyxl
        workbook1 = openpyxl.load_workbook(file1_path)
        workbook2 = openpyxl.load_workbook(file2_path)
        # Получаем активные листы в обоих файлах
        sheet1 = workbook1.active
        sheet2 = workbook2.active
    except Exception as e:
        # Обрабатываем возможные ошибки при открытии файлов
        result_label.config(text=f"Ошибка при открытии файлов: {str(e)}")
        return

    count = 0  # Счетчик обработанных записей кол-ва
    a1 = "Ремонт мебели и аксессуаров интерьера"
    a2 = "Монтаж мебели и аксессуаров интерьера"
    count1 = 0  # # Счетчик обработанных записей работ
    # Перебираем строки в первом файле
    for row1 in sheet1.iter_rows(min_row=2, max_col=12):
        cell1_col1 = row1[0].value  # Значение первого столбца
        cell1_col6 = row1[5].value  # Значение шестого столбца
        cell_col9 = row1[8].value  # Значение в 9 столбце
        cell_col12 = row1[11].value  # Значение в 12 столбце
        # Проверяем условия для строки в первом файле
        if cell1_col1 == "4600007963" and (cell1_col6 == 'В работе' or cell1_col6 == 'Новая') and cell_col12 is None:
            cell1_col3 = row1[2].value  # Значение третьего столбца

            # Перебираем строки во втором файле для поиска совпадения по значению третьего столбца
            for row2 in sheet2.iter_rows(min_row=2, max_col=6):
                cell2_col2 = row2[1].value  # Значение второго столбца во втором файле

                if cell2_col2 == cell1_col3:
                    cell2_col6 = row2[5].value  # Значение шестого столбца во втором файле
                    volume_start = cell2_col6.find("Объем работ: ")  # Находим начало "Объема работ"

                    if volume_start != -1:
                        volume_start += len("Объем работ: ")  # Увеличиваем индекс до конца "Объема работ"
                        volume_end = cell2_col6.find("\r", volume_start)  # Находим конец объема работ

                        if volume_end != -1:
                            volume = cell2_col6[volume_start:volume_end]  # Извлекаем значение объема работ
                            volume = volume.strip()  # Удаляем возможные пробелы по краям

                            try:
                                volume = float(volume)  # Преобразуем в вещественное число
                                row1[11].value = volume  # Записываем в первый файл
                                count += 1  # Увеличиваем счетчик
                            except ValueError as e:
                                # Обрабатываем ошибку преобразования строки в число
                                result_label.config(text=f"Ошибка преобразования в число: {str(e)}")

                            break  # Прекращаем поиск
        elif cell_col9 == a1 or cell_col9 == a2:
            cell2_col2 = row2[1].value  # Значение второго столбца во втором файле
            for row2 in sheet2.iter_rows(min_row=2, max_col=6):
                cell2_col2 = row2[1].value  # Значение второго столбца во втором файле

                if cell2_col2 == cell1_col3:
                    cell2_col6 = row2[5].value  # Значение шестого столбца во втором файле
                    volume_start = cell2_col6.find("Что требуется:")  # Находим начало "Что требуется:"

                    if volume_start != -1:
                        volume_start += len("Что требуется:")  # Увеличиваем индекс до конца "Что требуется:"
                        volume_end = cell2_col6.find("\r", volume_start)  # Находим конец Что требуется:

                        if volume_end != -1:
                            volume = cell2_col6[volume_start:volume_end]  # Извлекаем значение Что требуется:
                            volume = volume.strip()  # Удаляем возможные пробелы по краям
                            row1[8].value = volume  # Записываем в первый файл
                            count1 += 1
                            # Сохраняем измененный первый файл
    workbook1.save(file1_path)
    result_label.config(text=f"Обработано записей кол-ва: {count}")  # Выводим количество обработанных записей
    result_label.config(text=f"Обработано записей работ: {count1}")

# Создаем графический интерфейс с использованием Tkinter
root = tk.Tk()
root.title("Обработка Excel файлов")

# Создаем и размещаем элементы интерфейса
file1_label = tk.Label(root, text="Выберите первый файл:")
file1_label.pack()

file1_entry = tk.Entry(root)
file1_entry.pack()

file1_button = tk.Button(root, text="Обзор", command=lambda: file1_entry.insert(0, filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])))
file1_button.pack()

file2_label = tk.Label(root, text="Выберите второй файл:")
file2_label.pack()

file2_entry = tk.Entry(root)
file2_entry.pack()

file2_button = tk.Button(root, text="Обзор", command=lambda: file2_entry.insert(0, filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])))
file2_button.pack()

process_button = tk.Button(root, text="Обработать файлы", command=process_excel_files)
process_button.pack()

result_label = tk.Label(root, text="")
result_label.pack()

# Запускаем главный цикл Tkinter
root.mainloop()