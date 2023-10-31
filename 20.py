from tkinter import Tk, filedialog
from openpyxl import load_workbook

# Создаем графическое окно
root = Tk()
root.withdraw()

# Открываем диалоговое окно для выбора файла
file_path = filedialog.askopenfilename(filetypes=[('Excel Files', '*.xlsx')])

# Загружаем выбранный файл
book = load_workbook(filename=file_path)
sheet = book.active

i1 = sheet.max_row + 1  # количетво строк
zn = "!Без ВСП"
p0 = ["Что требуется:", "Вид неисправности:", "Вид нарушения работы:", "Что случилось:"]
p1 = "№"
p2 = "Мира 32к1"
p3 = "Алексея Кузнецова 38"  # 8629_01867
p4 = "Мира 44/20"
p5 = "\n"
p6 = "Ремонт, замена мебели и предметов интерьера"
p7 = "Монтаж/демонтаж мебели и предметов интерьера"
z1 = "Здание"
z2 = "8629_1867"
z3 = "Гараж"
obekt = ['г Великий Новгород  пр-кт Мира 32к1',
         'г Великий Новгород  пр-кт Мира 44/20',
         'г Боровичи  ул Алексея Кузнецова 38',
         'ст Уторгош ул Пионерская 65']
nomer_vsp = ['Здание', 'Гараж', '8629_01867', '8629_01532']

# Поиск и запись в ячейку CD выполняемой работы
for i in range(2, i1):
    b = sheet['bs' + str(i)].value  # содержимое ячейки bs
    i1 = 0
    while i1 < len(p0):  # применяем индекс для получения элемента
        index1 = b.find(p0[i1])
        if index1 == -1:
            pass
        else:
            Any = b.find(p5, index1)  #
            print(i, sheet['a' + str(i)].value, b[index1 + len(p0[i1]) + 1: Any])
            sheet['cd' + str(i)] = b[index1 + len(p0[i1]) + 1: Any]
        i1 += 1
    index6 = b.find('Здание:')
    index7 = b.find(p5, index6)
    index2 = b.find('№', index6, index7)
    if index2 == -1:
        i2 = 0
        while i2 < len(obekt):
            index4 = b.find(obekt[i2])
            if index4 != -1:
                sheet['cf' + str(i)] = nomer_vsp[i2]
                sheet['cg' + str(i)] = obekt[i2]
            else:
                pass
            i2 += 1
        # print(i, sheet['a' + str(i)].value, b[index4 + 15: 77:1])
    else:
        sheet['cf' + str(i)] = b[index2 + 2: index2 + 12:1]
        print(i, sheet['a' + str(i)].value, b[index2 + 2: index2 + 12: 1])
        sheet['cg' + str(i)] = b[index2 + 13: b.find(p5, index2): 1]
        print(i, sheet['a' + str(i)].value, b[index2 + 12: b.find(p5, index2): 1])

    if sheet['bf' + str(i)].value == p6:
        # print(sheet['bf' + str(i)].value)
        sheet['bf' + str(i)].value = sheet['cd' + str(i)].value
        # print(sheet['bf' + str(i)].value)
    if sheet['bf' + str(i)].value == p7:
        # print(sheet['bf' + str(i)].value)
        sheet['bf' + str(i)].value = sheet['cd' + str(i)].value
        # print(sheet['bf' + str(i)].value)

    # Проверка номера ВСП
    a = sheet['ae' + str(i)].value  # номер ВСП
    if a == zn:  # Проверка номера ВСП
        index = b.find(p1)
        if index != -1:
            # print(i, sheet['a' + str(i)].value, b[index + 2:index + 12])
            sheet['ae' + str(i)] = b[index + 2:index + 12]
        elif index != b.find(p2):
            # print(i, sheet['a' + str(i)].value, z1)
            sheet['ae' + str(i)] = z1
        elif index != b.find(p3):
            # print(i, sheet['a' + str(i)].value, z2)
            sheet['ae' + str(i)] = z2
        elif index != b.find(p4):
            # print(i, sheet['a' + str(i)].value, z3)
            sheet['ae' + str(i)] = z3

book.save(file_path)
book.close()
print("Файл обработан")